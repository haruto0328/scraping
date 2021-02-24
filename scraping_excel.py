import requests
from bs4 import BeautifulSoup
from urllib.parse import urljoin
import time

page_num = 1
tmp_num = 1
num = []
titles = []

while page_num <= 3:

    if page_num == 1:
        url = 'https://produce-web.net'
    else:
        url = urljoin(url, '/page/' + str(page_num))

    html = requests.get(url)
    soup = BeautifulSoup(html.content, "html.parser")

    for element in soup.select('.entry-title'):
        num.append(tmp_num)
        titles.append(element.text)
        tmp_num += 1

    page_num += 1
    time.sleep(1)

import openpyxl

wb = openpyxl.Workbook()
sheet = wb.active
sheet.title = 'blog_title' #excelシート名を「blog_title」に

sheet["A1"].value = '新着順'
sheet["B1"].value = '記事タイトル'

for i in range(1, tmp_num):
    sheet.cell(column=1, row=i+1, value=num[i-1])
    sheet.cell(column=2, row=i+1, value=titles[i-1])

wb.save('scraping_excel.xlsx')
wb.close()